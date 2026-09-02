# Install first: pip install openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

filename = "employees.xlsx"
workbook = Workbook()
sheet = workbook.active
sheet.title = "Employees"

# Add rows using append()
sheet.append(["ID", "Name", "Department", "Salary"])
sheet.append([101, "Asha", "IT", 65000])
sheet.append([102, "Ravi", "HR", 58000])
sheet.append([103, "Meena", "Finance", 72000])

# Access and modify cells
sheet["F1"] = "Total Payroll"
sheet["F2"] = "=SUM(D2:D4)"
sheet.cell(row=1, column=7, value="Employee Count")
sheet.cell(row=2, column=7, value="=COUNTA(B2:B4)")

# Format the header
for cell in sheet[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal="center")

# Adjust width, freeze header and add filtering
for column, width in {"A": 10, "B": 18, "C": 18, "D": 14}.items():
    sheet.column_dimensions[column].width = width
sheet.freeze_panes = "A2"
sheet.auto_filter.ref = "A1:D4"

# Add a second worksheet
summary_sheet = workbook.create_sheet("Summary")
summary_sheet.append(["Department", "Employees"])
summary_sheet.append(["IT", 1])
summary_sheet.append(["HR", 1])
summary_sheet.append(["Finance", 1])
workbook.save(filename)

# Load and read an existing workbook
loaded = load_workbook(filename)
print("Worksheet names:", loaded.sheetnames)
employee_sheet = loaded["Employees"]
print("Maximum row:", employee_sheet.max_row)
print("Maximum column:", employee_sheet.max_column)
print("Cell B2:", employee_sheet["B2"].value)

for row in employee_sheet.iter_rows(min_row=2, max_col=4, values_only=True):
    print(row)
